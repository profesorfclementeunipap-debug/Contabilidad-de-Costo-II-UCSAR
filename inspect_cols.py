import openpyxl

wb = openpyxl.load_workbook('notas_original.xlsx')
sheet = wb.active

for col in range(1, 20):
    val = sheet.cell(row=1, column=col).value
    val2 = sheet.cell(row=2, column=col).value
    val12 = sheet.cell(row=12, column=col).value
    if val:
        print(f"Col {col} (Row 1): {val}")
    if val2:
        print(f"Col {col} (Row 2): {val2}")
    if val12:
        print(f"Col {col} (Row 12): {val12}")
