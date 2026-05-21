import openpyxl

wb = openpyxl.load_workbook('notas_original.xlsx')
sheet = wb.active

# 1. Update header text
header_text = """Arquidiócesis de Caracas
Pontificia Universidad Católica Santa Rosa
“La Universidad del Diálogo y la Paz”
Escuela de Contaduría Pública
Período Académico 2026-II
Inicio: 22-05-2026 - Culmina: 31/07/2026"""
sheet.cell(row=1, column=1).value = header_text

# 2. Update PERIODO
sheet.cell(row=9, column=1).value = "PERIODO: 2026-II"

# 3. Update SUBJECT
sheet.cell(row=10, column=3).value = "CONTABILIDAD DE COSTO II"

# Save as new file
wb.save('Control de Notas Contabilidad de Costo II 2026-II.xlsx')
print("File saved successfully.")
